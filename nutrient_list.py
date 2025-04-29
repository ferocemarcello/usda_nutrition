import requests
import openpyxl
import time
import json
import os
nutrients = {203:"Protein", 204:"Total lipid (fat)", 205:"Carbohydrate, by difference", 208:"Energy",269:"Total Sugars",291:"Fiber, total dietary",601:"Cholesterol",606:"Fatty acids, total saturated",645:"Fatty acids, total monounsaturated",646:"Fatty acids, total polyunsaturated",605:"Fatty acids, total trans",210:"Sucrose",211:"Glucose",214:"Maltose",212:"Fructose",213:"Lactose",287:"Galactose",957:"Energy (Atwater General Factors)",958:"Energy (Atwater Specific Factors)",269.3:"Sugars, Total",298:"Total fat (NLEA)",693:"Fatty acids, total trans-monoenoic",695:"Fatty acids, total trans-polyenoic",205.2:"Carbohydrate, by summation",293:"Total dietary fiber (AOAC 2011.25)"}
nutrients_descriptions = list(nutrients.values())
nutrients_numbers = list(nutrients.keys())
FOOD_NUTRIENTS_KEY = "foodNutrients"
def fetch_food_list(api_key, page_number=1, page_size=100, max_retries=3):
    """Fetches a list of foods from the USDA FoodData Central API."""
    url = f"https://api.nal.usda.gov/fdc/v1/foods/list?api_key={api_key}&pageNumber={page_number}&pageSize={page_size}"
    retries = 0
    while retries < max_retries:
        try:
            response = requests.get(url)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 500:
                try:
                    error_json = response.json()
                    if "message" in error_json and "all shards failed" in error_json["message"]:
                        return "all_shards_failed"
                except json.JSONDecodeError:
                    pass
                retries += 1
                wait_time = 2 ** retries
                print(f"Server error (500) fetching food list (page {page_number}). Retrying in {wait_time} seconds...")
                time.sleep(wait_time)
            else:
                raise
        except requests.exceptions.RequestException as e:
            print(f"Error fetching food list (page {page_number}): {e}")
            return None
        return response.json()
    print(f"Max retries reached for fetching food list (page {page_number}).")
    return None

def fetch_food_details(api_key, fdc_id, nutrient_ids, max_retries=3):
    """Fetches detailed nutrient information for a specific food."""
    url = f"https://api.nal.usda.gov/fdc/v1/food/{fdc_id}?api_key={api_key}&nutrients={','.join(map(str, nutrient_ids))}"
    retries = 0
    while retries < max_retries:
        try:
            response = requests.get(url)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 404:
                print(f"Food with FDC ID {fdc_id} not found.")
                return None
            elif e.response.status_code == 500:
                retries += 1
                wait_time = 2 ** retries
                print(f"Server error (500) fetching details for FDC ID {fdc_id}. Retrying in {wait_time} seconds...")
                time.sleep(wait_time)
            else:
                raise
        except requests.exceptions.RequestException as e:
            print(f"Error fetching details for FDC ID {fdc_id}: {e}")
            return None
        return response.json()
    print(f"Max retries reached for fetching details for FDC ID {fdc_id}.")
    return None

def main_method(api_key, nutrient_ids, output_filename="food_nutrition_data.xlsx"):
    """Fetches food data and nutrient information and saves it to an Excel file.
    Only processes foods with FDC IDs greater than start_fdc_id.  Appends to existing file."""

    page_number = 1
    progress_counter = 0
    save_interval = 250  # Save progress every N foods
    workbook = None
    sheet = None
    file_exists = os.path.exists(output_filename)

    if file_exists:
        try:
            workbook = openpyxl.load_workbook(output_filename)
            sheet = workbook.active
        except Exception as e:
            print(f"Error opening existing workbook: {e}. Creating a new one.")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["fdcId", "description"] + nutrients_descriptions)  # Add header if creating new
            file_exists = False
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["fdcId", "description"] + nutrients_descriptions)  # Add header if creating new
    all_food_list = []
    while True:
        food_list_response:dict = fetch_food_list(api_key, page_number)
        if isinstance(food_list_response, str) and str.__eq__(food_list_response, "all_shards_failed"):
            break
        all_food_list.extend(food_list_response)
        page_number += 1
        time.sleep(0.1)
    print(f"len(all_food_list): {len(all_food_list)}")
    all_food_list = sorted(all_food_list, key=lambda item: item["description"])
    for food in all_food_list:
        fdc_id = food.get("fdcId")
        if fdc_id:
            description = food.get("description", "N/A")
            food_details = fetch_food_details(api_key, fdc_id, nutrient_ids)
            if food_details and FOOD_NUTRIENTS_KEY in food_details and food_details[FOOD_NUTRIENTS_KEY]:
                row_data = [fdc_id, description]
                food_values = food_details[FOOD_NUTRIENTS_KEY]
                try:
                    food_nutrients = [{"nutrient": x["nutrient"], "amount": x["amount"]} for x in food_values]
                except:
                    break
                nutrient_list = []
                for nutrient in food_nutrients:
                    nutrient_number = float(nutrient["nutrient"]["number"])
                    if nutrient_number in nutrient_ids:
                        nutrient_list.append({'nutrient number': nutrient_number, 'nutrient name': nutrient["nutrient"]["name"], 'unit_name':nutrient["nutrient"]["unitName"], 'amount': nutrient["amount"]})
                if nutrient_list:
                    nutrient_vals_to_append = []
                    nut_dict = {}
                    for x in nutrient_list:
                        nut_dict[x["nutrient number"]] = x["amount"]
                    for x in nutrients_numbers:
                        if x in nut_dict.keys():
                            nutrient_vals_to_append.append(nut_dict[x])
                        else:
                            nutrient_vals_to_append.append("N/A")
                    row_data.extend(nutrient_vals_to_append)
                    sheet.append(row_data)
                progress_counter += 1
                if progress_counter % save_interval == 0:
                    print(f"Processed {progress_counter} foods. Saving progress...")
                    try:
                        workbook.save(output_filename)
                    except Exception as e:
                        print(f"Error saving workbook: {e}")

    try:
        workbook.save(output_filename)
        print(f"Successfully saved data for {progress_counter} foods to {output_filename}")
    except Exception as e:
        print(f"Error saving final workbook: {e}")

if __name__ == "__main__":
    api_key = ""
    # Replace with the list of nutrient IDs you retrieved
    main_method(api_key, nutrients_numbers)
