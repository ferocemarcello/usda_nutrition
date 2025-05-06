import requests
import openpyxl
import os
import time
import json

nutrients = {
    203: "Protein",
    204: "Total lipid (fat)",
    205: "Carbohydrate, by difference",
    208: "Energy",
    269: "Total Sugars",
    291: "Fiber, total dietary",
    601: "Cholesterol",
    606: "Fatty acids, total saturated",
    645: "Fatty acids, total monounsaturated",
    646: "Fatty acids, total polyunsaturated",
    605: "Fatty acids, total trans",
    210: "Sucrose",
    211: "Glucose",
    214: "Maltose",
    212: "Fructose",
    213: "Lactose",
    287: "Galactose",
    957: "Energy (Atwater General Factors)",
    958: "Energy (Atwater Specific Factors)",
    269.3: "Sugars, Total",
    298: "Total fat (NLEA)",
    693: "Fatty acids, total trans-monoenoic",
    695: "Fatty acids, total trans-polyenoic",
    205.2: "Carbohydrate, by summation",
    293: "Total dietary fiber (AOAC 2011.25)",
}
nutrients_descriptions = list(nutrients.values())
nutrients_numbers = list(nutrients.keys())
FOOD_NUTRIENTS_KEY = "foodNutrients"
NOT_FOUND_IDS_FILENAME = "not_found_food_ids.json"
OUTPUT_FILENAME = "food_nutrition_data.xlsx"


def append_not_found_id(fdc_id):
    """Appends a not found FDC ID to a JSON file. Creates the file if it doesn't exist."""
    try:
        with open(NOT_FOUND_IDS_FILENAME, "r") as f:
            not_found_ids = json.load(f)
    except FileNotFoundError:
        not_found_ids = []
    except json.JSONDecodeError:
        print(
            f"Warning: Could not decode existing '{NOT_FOUND_IDS_FILENAME}'. Starting with an empty list."
        )
        not_found_ids = []

    if fdc_id not in not_found_ids:
        not_found_ids.append(fdc_id)
        with open(NOT_FOUND_IDS_FILENAME, "w") as f:
            json.dump(not_found_ids, f, indent=4)
        print(f"FDC ID {fdc_id} appended to '{NOT_FOUND_IDS_FILENAME}'.")


def fetch_food_details(api_key, fdc_id, nutrient_ids, max_retries=3):
    """Fetches detailed nutrient information for a specific food."""
    url = f"https://api.nal.usda.gov/fdc/v1/food/{fdc_id}?api_key={api_key}&nutrients={','.join(map(str, nutrient_ids))}"
    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 404:
            print(f"Food with FDC ID {fdc_id} not found.")
            append_not_found_id(fdc_id)
            return None
        elif e.response.status_code == 500:
            print(
                f"Server error (500) fetching details for FDC ID {fdc_id}."
            )
            append_not_found_id(fdc_id)
            return None
        else:
            raise
    except requests.exceptions.RequestException as e:
        print(f"Error fetching details for FDC ID {fdc_id}: {e}")
        append_not_found_id(fdc_id)
        return None


def get_existing_fdc_ids(filename=OUTPUT_FILENAME):
    """Fetches a list of fdcId from the specified Excel file."""
    existing_ids = set()
    if os.path.exists(filename):
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            # Assuming 'fdcId' is in the first column (index 0)
            for row in sheet:
                if row and row[0].value is not None and str(row[0].value).isdigit():
                    existing_ids.add(int(row[0].value))
            print(f"Found {len(existing_ids)} existing FDC IDs in '{filename}'.")
        except Exception as e:
            print(f"Error reading existing Excel file '{filename}': {e}")
    else:
        print(f"Excel file '{filename}' not found. Starting fresh.")
    return existing_ids


def main_method(
    api_key,
    nutrient_ids,
    json_filename="food_ids.json",
    output_filename=OUTPUT_FILENAME,
    amount_foods_to_process: int = None,
):
    """Fetches food data and nutrient information from a JSON file and saves it to an Excel file."""

    existing_fdc_ids = get_existing_fdc_ids(output_filename)
    progress_counter = 0
    save_interval = 250  # Save progress every N foods
    workbook = None
    sheet = None
    file_exists = os.path.exists(output_filename)

    if file_exists:
        try:
            workbook = openpyxl.load_workbook(output_filename)
            sheet = workbook.active
        except Exception as e:
            print(f"Error opening existing workbook: {e}. Creating a new one.")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(
                ["fdcId", "description"] + nutrients_descriptions
            )  # Add header if creating new
            file_exists = False
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(
            ["fdcId", "description"] + nutrients_descriptions
        )  # Add header if creating new

    try:
        with open(json_filename, "r", encoding="utf-8") as f:
            food_data = json.load(f)
    except FileNotFoundError:
        print(f"Error: The file '{json_filename}' was not found.")
        return
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON in file '{json_filename}'.")
        return

    filtered_food_data = [
        food for food in food_data if food.get("fdcId") not in existing_fdc_ids
    ]
    
    print(
        f"Using {len(filtered_food_data)} new foods (out of {len(food_data)} in JSON)."
    )

    filtered_food_data = filtered_food_data[:amount_foods_to_process]

    print(
        f"Processing {len(filtered_food_data)} new foods (out of {len(food_data)} in JSON)."
    )
    for food in filtered_food_data:
        fdc_id = food.get("fdcId")
        description = food.get("description", "N/A")
        if fdc_id:
            food_details = fetch_food_details(api_key, fdc_id, nutrient_ids)
            if (
                food_details
                and FOOD_NUTRIENTS_KEY in food_details
                and food_details[FOOD_NUTRIENTS_KEY]
            ):
                row_data = [fdc_id, description]
                food_values = food_details[FOOD_NUTRIENTS_KEY]
                food_nutrients = [
                    (
                        {"nutrient": x["nutrient"], "amount": x["amount"]}
                        if "amount" in x.keys()
                        else {"nutrient": x["nutrient"], "amount": "N/A"}
                    )
                    for x in food_values
                ]
                nutrient_list = []
                for nutrient in food_nutrients:
                    nutrient_number = float(nutrient["nutrient"]["number"])
                    if nutrient_number in nutrient_ids:
                        nutrient_list.append(
                            {
                                "nutrient number": nutrient_number,
                                "nutrient name": nutrient["nutrient"]["name"],
                                "unit_name": nutrient["nutrient"]["unitName"],
                                "amount": nutrient["amount"],
                            }
                        )
                if nutrient_list:
                    nutrient_vals_to_append = []
                    nut_dict = {}
                    for x in nutrient_list:
                        nut_dict[x["nutrient number"]] = x["amount"]
                    for x in nutrients_numbers:
                        if x in nut_dict.keys():
                            nutrient_vals_to_append.append(nut_dict[x])
                        else:
                            nutrient_vals_to_append.append("N/A")
                    row_data.extend(nutrient_vals_to_append)
                    sheet.append(row_data)
                    progress_counter += 1
                    if progress_counter % save_interval == 0:
                        print(
                            f"Processed {progress_counter} new foods. Saving progress..."
                        )
                        try:
                            workbook.save(output_filename)
                        except Exception as e:
                            print(f"Error saving workbook: {e}")

    try:
        workbook.save(output_filename)
        print(
            f"Successfully saved data for {progress_counter} new foods to {output_filename}"
        )
    except Exception as e:
        print(f"Error saving final workbook: {e}")


if __name__ == "__main__":
    api_key = ""
    json_filename = "food_ids.json"  # Specify the name of your JSON file
    main_method(api_key, nutrients_numbers, json_filename, amount_foods_to_process=2000)
